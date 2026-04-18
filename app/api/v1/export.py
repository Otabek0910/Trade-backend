from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import cast, Date, func, text
from datetime import date, timedelta
from io import BytesIO
import httpx
import subprocess
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.models.sale import Sale, SaleStatus
from app.models.sale_item import SaleItem
from app.models.product import Product
from app.models.customer import Customer
from app.models.supplier import Supplier
from app.models.supplier_return import SupplierReturn
from app.models.supplier_payment import SupplierPayment
from app.models.receipt import Receipt
from app.models.user import User, UserRole
from app.models.expense import Expense
from app.models.return_model import Return
from app.core.telegram_auth import get_current_user

router = APIRouter(prefix="/export", tags=["Экспорт"])

# ─── Стили ───────────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", start_color="1A4B8C")
HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEADER_FILL = PatternFill("solid", start_color="E8F0FE")
SUBHEADER_FONT = Font(bold=True, color="1A4B8C", name="Arial", size=10)
NORMAL_FONT    = Font(name="Arial", size=10)
RED_FONT       = Font(name="Arial", size=10, color="C0392B", bold=True)
GREEN_FONT     = Font(name="Arial", size=10, color="1A6B3C", bold=True)
ORANGE_FONT    = Font(name="Arial", size=10, color="E08030", bold=True)
WARN_FILL      = PatternFill("solid", start_color="FFF3CD")
RED_FILL       = PatternFill("solid", start_color="FDECEA")
GREEN_FILL     = PatternFill("solid", start_color="E8F5E9")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="1A4B8C")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, values, widths=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border()
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

def data_cell(ws, row, col, value, number_format=None, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or NORMAL_FONT
    cell.border = thin_border()
    cell.alignment = align or LEFT
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill
    return cell

def kpi_block(ws, row, col, label, value, sub=None, label_color="1A1A1A", value_color="1A4B8C"):
    """KPI блок: метка + значение + подпись"""
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name="Arial", size=9, color="888888")
    lc.alignment = LEFT

    vc = ws.cell(row=row+1, column=col, value=value)
    vc.font = Font(name="Arial", size=13, bold=True, color=value_color)
    vc.alignment = LEFT
    vc.number_format = '#,##0'

    if sub:
        sc = ws.cell(row=row+2, column=col, value=sub)
        sc.font = Font(name="Arial", size=8, color="888888", italic=True)
        sc.alignment = LEFT


# ─── Лист 0: Сводка (Summary) ─────────────────────────────────────────────────

def build_summary_sheet(ws, db: Session, date_from: date, date_to: date):
    ws.title = "📊 Сводка"
    ws.sheet_view.showGridLines = False

    # ── Сбор данных ──────────────────────────────────────────────────────────
    completed_sales = db.query(Sale).filter(
        cast(Sale.created_at, Date) >= date_from,
        cast(Sale.created_at, Date) <= date_to,
        Sale.status == SaleStatus.completed,
    ).all()
    returned_sales = db.query(Sale).filter(
        cast(Sale.created_at, Date) >= date_from,
        cast(Sale.created_at, Date) <= date_to,
        Sale.status == SaleStatus.returned,
    ).all()

    all_sales_in_period = completed_sales + returned_sales
    revenue_gross  = sum(float(s.total_amount) for s in all_sales_in_period)
    paid_received  = sum(float(s.paid_amount)  for s in completed_sales)

    cogs = sum(float(i.purchase_price_at_sale)*i.quantity for s in all_sales_in_period for i in s.items)

    returns_in_period = db.query(Return).join(Sale, Return.sale_id==Sale.id).filter(
        cast(Return.created_at, Date) >= date_from,
        cast(Return.created_at, Date) <= date_to,
    ).all()
    returns_amount = sum(float(r.return_amount) for r in returns_in_period)
    returns_cogs = 0.0
    for r in returns_in_period:
        si = db.query(SaleItem).filter(SaleItem.sale_id==r.sale_id, SaleItem.product_id==r.product_id).first()
        if si: returns_cogs += float(si.purchase_price_at_sale)*r.quantity

    net_revenue  = revenue_gross - returns_amount
    net_cogs     = cogs - returns_cogs
    gross_profit = net_revenue - net_cogs
    margin_pct   = round(gross_profit/net_revenue*100, 1) if net_revenue > 0 else 0

    expenses_total = float(db.query(func.sum(Expense.amount)).filter(
        Expense.date >= date_from, Expense.date <= date_to
    ).scalar() or 0)
    net_profit = gross_profit - expenses_total

    total_cust_debt = float(db.query(func.sum(Customer.total_debt)).scalar() or 0)
    total_sup_debt  = float(db.query(func.sum(Supplier.total_debt)).scalar() or 0)
    cash_in_hand    = paid_received - max(0, returns_amount - total_cust_debt)
    cash_pct        = round(cash_in_hand/net_revenue*100, 1) if net_revenue > 0 else 0

    total_purchased_raw = float(db.query(func.sum(Receipt.paid_amount + Receipt.debt)).scalar() or 0)
    total_returned_raw  = float(db.query(func.sum(SupplierReturn.return_amount)).scalar() or 0)
    total_purchased_net = total_purchased_raw - total_returned_raw

    stock_value     = float(db.query(func.sum(Product.purchase_price*Product.current_stock)).scalar() or 0)
    stock_potential = float(db.query(func.sum(Product.selling_price *Product.current_stock)).scalar() or 0)

    sales_count    = len(completed_sales)
    returned_count = len(returned_sales)
    returns_count  = len(returns_in_period)

    top_debtors = db.query(Customer).filter(Customer.total_debt>0).order_by(Customer.total_debt.desc()).limit(8).all()

    # ── Шапка ─────────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 2
    for col_letter, w in zip(['B','C','D','E','F','G','H','I','J'], [28,16,16,16,16,14,14,14,14]):
        ws.column_dimensions[col_letter].width = w

    # Заголовок
    ws.merge_cells("B1:J1")
    t = ws["B1"]
    t.value = "TRADI · Отчёт по продажам"
    t.font = Font(name="Arial", size=16, bold=True, color="1A4B8C")
    t.alignment = LEFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("B2:J2")
    sub = ws["B2"]
    period_label = "Вся история" if date_from.year == 2000 else f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    sub.value = f"Период: {period_label}   |   Сформировано: {date.today().strftime('%d.%m.%Y')}"
    sub.font = Font(name="Arial", size=10, color="888888")
    sub.alignment = LEFT
    ws.row_dimensions[2].height = 18

    # ── KPI строка ─────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 14

    kpi_items = [
        ("ВЛОЖИЛИ", total_purchased_net, "закупки (нетто)", "B", "7A3B8C"),
        ("ЧИСТАЯ ВЫРУЧКА", net_revenue, f"после возвратов −{returns_amount:,.0f}", "D", "1A6B3C"),
        ("В КАССЕ РЕАЛЬНО", cash_in_hand, f"{cash_pct}% от выручки", "F", "2481CC"),
        ("ЧИСТАЯ ПРИБЫЛЬ", net_profit, f"маржа {margin_pct}%", "H", "34A853" if net_profit >= 0 else "C0392B"),
    ]
    for label, value, sub_text, col, color in kpi_items:
        lc = ws[f"{col}4"]
        lc.value = label
        lc.font = Font(name="Arial", size=8, color="888888", bold=True)

        vc = ws[f"{col}5"]
        vc.value = round(value)
        vc.font = Font(name="Arial", size=14, bold=True, color=color)
        vc.number_format = '#,##0'

        sc = ws[f"{col}6"]
        sc.value = sub_text
        sc.font = Font(name="Arial", size=8, color="AAAAAA", italic=True)

    # Долг поставщику — отдельно справа
    ws["J4"].value = "ДОЛГ ПОСТАВЩИКУ"
    ws["J4"].font = Font(name="Arial", size=8, color="888888", bold=True)
    ws["J5"].value = round(total_sup_debt)
    ws["J5"].font = Font(name="Arial", size=14, bold=True, color="C0392B")
    ws["J5"].number_format = '#,##0'
    ws["J6"].value = "не оплачено"
    ws["J6"].font = Font(name="Arial", size=8, color="AAAAAA", italic=True)

    # Разделитель
    ws.row_dimensions[7].height = 6

    # ── Левый блок: Закупки + Склад ────────────────────────────────────────────
    row = 8

    # Закупки
    ws.merge_cells(f"B{row}:E{row}")
    h = ws[f"B{row}"]
    h.value = "ЗАКУПКИ У ПОСТАВЩИКА"
    h.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", start_color="1A4B8C")
    h.alignment = LEFT
    ws.row_dimensions[row].height = 18

    zakupki_rows = [
        ("Куплено товара всего (нетто)", total_purchased_net, "1A1A1A"),
        ("Уже оплатили", total_purchased_net - total_sup_debt, "1A6B3C"),
        ("Ещё должны поставщику", total_sup_debt, "C0392B"),
    ]
    for i, (lbl, val, color) in enumerate(zakupki_rows):
        r2 = row + 1 + i
        ws.merge_cells(f"B{r2}:D{r2}")
        lc = ws[f"B{r2}"]
        lc.value = lbl
        lc.font = NORMAL_FONT
        lc.border = thin_border()
        lc.alignment = LEFT

        vc = ws[f"E{r2}"]
        vc.value = round(val)
        vc.font = Font(name="Arial", size=10, bold=True, color=color)
        vc.number_format = '#,##0'
        vc.alignment = RIGHT
        vc.border = thin_border()
        if color == "C0392B":
            vc.fill = PatternFill("solid", start_color="FDECEA")
        ws.row_dimensions[r2].height = 18

    row += 4

    # Склад
    ws.merge_cells(f"B{row}:E{row}")
    h2 = ws[f"B{row}"]
    h2.value = "ТОВАР НА СКЛАДЕ СЕЙЧАС"
    h2.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    h2.fill = PatternFill("solid", start_color="1A6B3C")
    h2.alignment = LEFT
    ws.row_dimensions[row].height = 18

    sklad_rows = [
        ("Стоимость по закупке", stock_value, "1A1A1A"),
        ("Потенциал по ценам продажи", stock_potential, "7A3B8C"),
    ]
    for i, (lbl, val, color) in enumerate(sklad_rows):
        r2 = row + 1 + i
        ws.merge_cells(f"B{r2}:D{r2}")
        lc = ws[f"B{r2}"]
        lc.value = lbl; lc.font = NORMAL_FONT; lc.border = thin_border(); lc.alignment = LEFT
        vc = ws[f"E{r2}"]
        vc.value = round(val)
        vc.font = Font(name="Arial", size=10, bold=True, color=color)
        vc.number_format = '#,##0'; vc.alignment = RIGHT; vc.border = thin_border()
        ws.row_dimensions[r2].height = 18

    row += 3

    # ── Правый блок: Выручка и прибыль ────────────────────────────────────────
    col_start = 7  # G
    r_right = 8

    ws.merge_cells(f"G{r_right}:J{r_right}")
    h3 = ws[f"G{r_right}"]
    h3.value = "ВЫРУЧКА И ПРИБЫЛЬ"
    h3.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    h3.fill = PatternFill("solid", start_color="1A6B3C")
    h3.alignment = LEFT
    ws.row_dimensions[r_right].height = 18

    pl_rows = [
        (f"Продажи ({sales_count} завершённых)", revenue_gross, False, "1A1A1A"),
        (f"Возвраты ({returns_count} шт)", -returns_amount, False, "E08030"),
        ("= Чистая выручка", net_revenue, True, "1A6B3C"),
        ("Минус себестоимость", -net_cogs, False, "555555"),
        (f"= Валовая прибыль  {margin_pct}%", gross_profit, True, "1A6B3C"),
        ("Минус расходы", -expenses_total, False, "C0392B"),
        ("Чистая прибыль", net_profit, True, "34A853" if net_profit >= 0 else "C0392B"),
    ]
    for i, (lbl, val, bold, color) in enumerate(pl_rows):
        r2 = r_right + 1 + i
        ws.merge_cells(f"G{r2}:I{r2}")
        lc = ws[f"G{r2}"]
        lc.value = lbl
        lc.font = Font(name="Arial", size=9, bold=bold)
        lc.border = thin_border()
        lc.alignment = LEFT
        if bold:
            lc.fill = PatternFill("solid", start_color="E8F0FE")

        vc = ws[f"J{r2}"]
        vc.value = round(val)
        vc.font = Font(name="Arial", size=9, bold=bold, color=color)
        vc.number_format = '#,##0'
        vc.alignment = RIGHT
        vc.border = thin_border()
        if bold:
            vc.fill = PatternFill("solid", start_color="E8F0FE")
        ws.row_dimensions[r2].height = 18

    r_right += len(pl_rows) + 1

    # Где деньги
    ws.merge_cells(f"G{r_right}:J{r_right}")
    hd = ws[f"G{r_right}"]
    hd.value = "ГДЕ СЕЙЧАС ДЕНЬГИ?"
    hd.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hd.fill = PatternFill("solid", start_color="C0392B")
    hd.alignment = LEFT
    ws.row_dimensions[r_right].height = 18

    money_rows = [
        ("Чистая выручка", net_revenue, False, "1A1A1A"),
        ("Клиенты не заплатили", -total_cust_debt, False, "C0392B"),
        (f"Фактически в кассе  ({cash_pct}%)", cash_in_hand, True, "2481CC"),
    ]
    for i, (lbl, val, bold, color) in enumerate(money_rows):
        r2 = r_right + 1 + i
        ws.merge_cells(f"G{r2}:I{r2}")
        lc = ws[f"G{r2}"]
        lc.value = lbl
        lc.font = Font(name="Arial", size=9, bold=bold)
        lc.border = thin_border()
        lc.alignment = LEFT
        if bold: lc.fill = GREEN_FILL
        vc = ws[f"J{r2}"]
        vc.value = round(val)
        vc.font = Font(name="Arial", size=9, bold=bold, color=color)
        vc.number_format = '#,##0'
        vc.alignment = RIGHT
        vc.border = thin_border()
        if bold: vc.fill = GREEN_FILL
        ws.row_dimensions[r2].height = 18

    # ── Должники ───────────────────────────────────────────────────────────────
    base_row = max(row, r_right + 5) + 2

    ws.merge_cells(f"B{base_row}:J{base_row}")
    hd2 = ws[f"B{base_row}"]
    hd2.value = f"ТОП ДОЛЖНИКОВ  (итого {total_cust_debt:,.0f} сум)"
    hd2.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hd2.fill = PatternFill("solid", start_color="C0392B")
    hd2.alignment = LEFT
    ws.row_dimensions[base_row].height = 18

    header_row(ws, base_row+1, ["#", "Клиент", "Телефон", "Долг (сум)", "% от покупок"])
    ws.column_dimensions['B'].width = 4
    for i, c in enumerate(top_debtors, 1):
        r2 = base_row + 1 + i
        pct = round(float(c.total_debt)/float(c.total_purchases or 1)*100, 1) if c.total_purchases else 0
        data_cell(ws, r2, 2, i, align=CENTER)
        data_cell(ws, r2, 3, c.name)
        data_cell(ws, r2, 4, c.phone or "—")
        data_cell(ws, r2, 5, float(c.total_debt), number_format='#,##0', align=RIGHT, font=RED_FONT,
                  fill=RED_FILL if float(c.total_debt) > 1_000_000 else None)
        data_cell(ws, r2, 6, pct, number_format='0.0"%"', align=CENTER)
        ws.row_dimensions[r2].height = 18

    # Итог
    end_row = base_row + 1 + len(top_debtors) + 1
    ws.merge_cells(f"B{end_row}:D{end_row}")
    ws.cell(end_row, 2).value = "ИТОГО ДОЛГ КЛИЕНТОВ"
    ws.cell(end_row, 2).font = SUBHEADER_FONT
    ws.cell(end_row, 2).fill = SUBHEADER_FILL
    ws.cell(end_row, 2).border = thin_border()
    ws.cell(end_row, 2).alignment = LEFT
    vc = ws.cell(end_row, 5)
    vc.value = round(total_cust_debt)
    vc.number_format = '#,##0'
    vc.font = RED_FONT; vc.fill = SUBHEADER_FILL; vc.border = thin_border(); vc.alignment = RIGHT
    ws.row_dimensions[end_row].height = 20

    # Финальный вывод
    final_row = end_row + 2
    ws.merge_cells(f"B{final_row}:J{final_row}")
    fc = ws[f"B{final_row}"]
    fc.value = (
        f"ГЛАВНЫЙ ВЫВОД: Из {net_revenue:,.0f} сум чистой выручки в кассе реально {cash_in_hand:,.0f} сум ({cash_pct}%). "
        f"Долг клиентов: {total_cust_debt:,.0f} сум. "
        f"Долг поставщику: {total_sup_debt:,.0f} сум. "
        f"Маржа: {margin_pct}%."
    )
    fc.font = Font(name="Arial", size=9, bold=True, color="1A4B8C")
    fc.fill = PatternFill("solid", start_color="E8F0FE")
    fc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fc.border = thick_border()
    ws.row_dimensions[final_row].height = 36


# ─── Лист 1: Продажи ─────────────────────────────────────────────────────────

def build_sales_sheet(ws, db: Session, date_from: date, date_to: date):
    ws.title = "Продажи"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"Продажи за период {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="1A4B8C")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["№", "Дата", "Статус", "Клиент", "Продавец", "Товар", "Кол-во", "Цена", "Сумма", "Оплачено", "Долг"]
    widths = [5, 16, 12, 22, 18, 30, 8, 16, 16, 16, 14]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    sales = db.query(Sale).filter(
        cast(Sale.created_at, Date) >= date_from,
        cast(Sale.created_at, Date) <= date_to,
        Sale.status.in_([SaleStatus.completed, SaleStatus.returned]),
    ).order_by(Sale.created_at.desc()).all()

    r = 3
    total_revenue = total_paid = 0

    for sale in sales:
        is_returned = sale.status == SaleStatus.returned
        for item in sale.items:
            debt       = float(sale.total_amount) - float(sale.paid_amount)
            item_share = debt / len(sale.items) if sale.items else 0
            item_total = float(item.selling_price) * item.quantity

            data_cell(ws, r, 1, sale.id, align=CENTER)
            data_cell(ws, r, 2, sale.created_at.strftime("%d.%m.%Y %H:%M") if sale.created_at else "")
            sc = data_cell(ws, r, 3, "↩ Возврат" if is_returned else "✅ Продажа")
            if is_returned: sc.font = ORANGE_FONT
            data_cell(ws, r, 4, sale.customer.name if sale.customer else "Розница")
            data_cell(ws, r, 5, sale.seller.full_name if sale.seller else "—")
            data_cell(ws, r, 6, item.product.name if item.product else "—")
            data_cell(ws, r, 7, item.quantity, align=CENTER)
            data_cell(ws, r, 8, float(item.selling_price), number_format='#,##0', align=RIGHT)
            data_cell(ws, r, 9, item_total, number_format='#,##0', align=RIGHT,
                      font=Font(name="Arial", size=10, color="888888") if is_returned else GREEN_FONT)
            data_cell(ws, r, 10, float(sale.paid_amount), number_format='#,##0', align=RIGHT)
            dc = data_cell(ws, r, 11, item_share if item_share > 0 else 0, number_format='#,##0', align=RIGHT)
            if item_share > 0: dc.font = RED_FONT
            ws.row_dimensions[r].height = 18

            if not is_returned:
                total_revenue += item_total
                total_paid    += float(sale.paid_amount) / len(sale.items)
            r += 1

    total_debt = total_revenue - total_paid
    ws.row_dimensions[r].height = 22
    for col in range(1, 12):
        ws.cell(r, col).fill = SUBHEADER_FILL
        ws.cell(r, col).border = thin_border()
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1).value = "ИТОГО"
    ws.cell(r, 1).font = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    data_cell(ws, r, 9,  total_revenue, number_format='#,##0', align=RIGHT, font=Font(bold=True, name="Arial", size=10, color="1A6B3C"), fill=SUBHEADER_FILL)
    data_cell(ws, r, 10, total_paid,    number_format='#,##0', align=RIGHT, font=Font(bold=True, name="Arial", size=10), fill=SUBHEADER_FILL)
    data_cell(ws, r, 11, total_debt,    number_format='#,##0', align=RIGHT, font=Font(bold=True, name="Arial", size=10, color="C0392B" if total_debt > 0 else "1A6B3C"), fill=SUBHEADER_FILL)


# ─── Лист 2: Остатки ─────────────────────────────────────────────────────────

def build_stock_sheet(ws, db: Session):
    ws.title = "Остатки"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = f"Остатки на складе по состоянию на {date.today().strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="1A6B3C")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["SKU", "Название", "Бренд", "Категория", "Поставщик", "Валюта", "Цена закупки", "Цена продажи", "Маржа %", "Остаток", "Мин.остаток", "Стоимость склада"]
    widths = [14, 30, 14, 16, 20, 8, 16, 16, 10, 10, 11, 18]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    products = db.query(Product).order_by(Product.name).all()

    r = 3
    total_stock_value = 0
    for p in products:
        margin = round((float(p.selling_price)-float(p.purchase_price))/float(p.purchase_price)*100, 1) if float(p.purchase_price) > 0 else 0
        low    = p.current_stock <= p.min_stock
        row_fill = RED_FILL if low else None
        stock_value = float(p.purchase_price) * p.current_stock

        data_cell(ws, r, 1,  p.sku,                                              fill=row_fill)
        data_cell(ws, r, 2,  p.name,                                             fill=row_fill)
        data_cell(ws, r, 3,  p.brand or "—",                                     fill=row_fill)
        data_cell(ws, r, 4,  p.category or "—",                                  fill=row_fill)
        data_cell(ws, r, 5,  p.supplier.name if p.supplier else "—",             fill=row_fill)
        data_cell(ws, r, 6,  (p.purchase_currency or "uzs").upper(),  align=CENTER, fill=row_fill)
        data_cell(ws, r, 7,  float(p.purchase_price), number_format='#,##0', align=RIGHT, fill=row_fill)
        data_cell(ws, r, 8,  float(p.selling_price),  number_format='#,##0', align=RIGHT, fill=row_fill)
        data_cell(ws, r, 9,  margin, number_format='0.0"%"', align=CENTER,
                  font=GREEN_FONT if margin >= 20 else NORMAL_FONT, fill=row_fill)
        sc = data_cell(ws, r, 10, p.current_stock, align=CENTER, fill=row_fill)
        if low: sc.font = RED_FONT
        data_cell(ws, r, 11, p.min_stock, align=CENTER, fill=row_fill)
        data_cell(ws, r, 12, stock_value, number_format='#,##0', align=RIGHT, fill=row_fill)
        ws.row_dimensions[r].height = 18
        total_stock_value += stock_value
        r += 1

    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:K{r}")
    ws.cell(r, 1).value = "ИТОГО СТОИМОСТЬ СКЛАДА"
    ws.cell(r, 1).font = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    ws.cell(r, 1).fill = SUBHEADER_FILL
    ws.cell(r, 1).border = thin_border()
    data_cell(ws, r, 12, total_stock_value, number_format='#,##0', align=RIGHT,
              font=Font(bold=True, name="Arial", size=10, color="1A6B3C"), fill=SUBHEADER_FILL)


# ─── Лист 3: Долги клиентов ──────────────────────────────────────────────────

def build_debts_sheet(ws, db: Session):
    ws.title = "Долги клиентов"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"Долги клиентов на {date.today().strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="C0392B")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["Клиент", "Телефон", "Адрес", "Всего покупок", "Долг", "% долга"]
    widths = [28, 18, 24, 18, 18, 14]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    debtors = db.query(Customer).filter(Customer.total_debt > 0).order_by(Customer.total_debt.desc()).all()

    r = 3
    for c in debtors:
        purchases = float(c.total_purchases or 0)
        debt      = float(c.total_debt)
        pct       = round(debt/purchases*100, 1) if purchases > 0 else 0
        fill      = RED_FILL if debt > 1_000_000 else None

        data_cell(ws, r, 1, c.name,     fill=fill)
        data_cell(ws, r, 2, c.phone or "—",  fill=fill)
        data_cell(ws, r, 3, c.address or "—", fill=fill)
        data_cell(ws, r, 4, purchases, number_format='#,##0', align=RIGHT, fill=fill)
        data_cell(ws, r, 5, debt,      number_format='#,##0', align=RIGHT, font=RED_FONT, fill=fill)
        data_cell(ws, r, 6, pct,       number_format='0.0"%"', align=CENTER,
                  font=RED_FONT if pct > 50 else NORMAL_FONT, fill=fill)
        ws.row_dimensions[r].height = 18
        r += 1

    if r == 3:
        ws.merge_cells("A3:F3")
        ws.cell(3, 1).value = "✅ Долгов нет"
        ws.cell(3, 1).font  = Font(name="Arial", size=11, color="1A6B3C", bold=True)
        ws.cell(3, 1).alignment = CENTER
        r = 4

    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(r, 1).value = "ИТОГО ДОЛГ"
    ws.cell(r, 1).font  = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    ws.cell(r, 1).fill  = SUBHEADER_FILL
    ws.cell(r, 1).border = thin_border()
    total_d = sum(float(c.total_debt) for c in debtors)
    data_cell(ws, r, 5, total_d, number_format='#,##0', align=RIGHT,
              font=RED_FONT, fill=SUBHEADER_FILL)


# ─── Лист 4: Поставщики ──────────────────────────────────────────────────────

def build_suppliers_sheet(ws, db: Session):
    ws.title = "Поставщики"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Поставщики и расчёты на {date.today().strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="7A3B8C")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["Поставщик", "Телефон", "Товаров", "Приёмок", "Закуплено (нетто)", "Оплачено", "Наш долг", "Нам должны", "Возвратов"]
    widths = [24, 16, 10, 10, 20, 20, 18, 16, 12]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    suppliers = db.query(Supplier).order_by(Supplier.name).all()

    r = 3
    for s in suppliers:
        prod_count  = db.query(func.count(Product.id)).filter(Product.supplier_id==s.id).scalar() or 0
        rec_count   = db.query(func.count(Receipt.id)).filter(Receipt.supplier_id==s.id).scalar() or 0
        purchased   = float(db.query(func.sum(Receipt.paid_amount+Receipt.debt)).filter(Receipt.supplier_id==s.id).scalar() or 0)
        returned    = float(db.query(func.sum(SupplierReturn.return_amount)).filter(SupplierReturn.supplier_id==s.id).scalar() or 0)
        paid_direct = float(db.query(func.sum(Receipt.paid_amount)).filter(Receipt.supplier_id==s.id).scalar() or 0)
        extra_paid  = float(db.query(func.sum(SupplierPayment.amount)).filter(SupplierPayment.supplier_id==s.id, ~SupplierPayment.note.like('%💚%')).scalar() or 0)
        net_purchased = purchased - returned
        total_paid  = paid_direct + extra_paid
        total_debt  = float(s.total_debt or 0)
        total_credit= float(getattr(s, 'total_credit', 0) or 0)
        ret_count   = db.query(func.count(SupplierReturn.id)).filter(SupplierReturn.supplier_id==s.id).scalar() or 0

        fill = RED_FILL if total_debt > 0 else (GREEN_FILL if total_credit > 0 else None)
        data_cell(ws, r, 1, s.name, fill=fill)
        data_cell(ws, r, 2, s.phone or "—", fill=fill)
        data_cell(ws, r, 3, prod_count, align=CENTER, fill=fill)
        data_cell(ws, r, 4, rec_count,  align=CENTER, fill=fill)
        data_cell(ws, r, 5, net_purchased, number_format='#,##0', align=RIGHT, fill=fill)
        data_cell(ws, r, 6, total_paid,    number_format='#,##0', align=RIGHT, fill=fill)
        data_cell(ws, r, 7, total_debt,    number_format='#,##0', align=RIGHT,
                  font=RED_FONT if total_debt > 0 else NORMAL_FONT, fill=fill)
        data_cell(ws, r, 8, total_credit,  number_format='#,##0', align=RIGHT,
                  font=GREEN_FONT if total_credit > 0 else NORMAL_FONT, fill=fill)
        data_cell(ws, r, 9, ret_count, align=CENTER, fill=fill)
        ws.row_dimensions[r].height = 18
        r += 1

    ws.row_dimensions[r].height = 22
    for col in range(1, 10):
        ws.cell(r, col).fill = SUBHEADER_FILL
        ws.cell(r, col).border = thin_border()
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(r, 1).value = "ИТОГО"
    ws.cell(r, 1).font  = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    all_debt   = float(db.query(func.sum(Supplier.total_debt)).scalar() or 0)
    all_credit = float(db.query(func.sum(Supplier.total_credit)).scalar() or 0)
    data_cell(ws, r, 7, all_debt,   number_format='#,##0', align=RIGHT, font=RED_FONT,   fill=SUBHEADER_FILL)
    data_cell(ws, r, 8, all_credit, number_format='#,##0', align=RIGHT, font=GREEN_FONT, fill=SUBHEADER_FILL)


# ─── Лист 5: Расходы ─────────────────────────────────────────────────────────

def build_expenses_sheet(ws, db: Session, date_from: date, date_to: date):
    ws.title = "Расходы"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"Расходы за период {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="C0392B")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["Дата", "Категория", "Описание", "Сумма", "Добавил"]
    widths = [16, 20, 36, 16, 22]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    expenses = db.query(Expense).filter(
        Expense.date >= date_from, Expense.date <= date_to
    ).order_by(Expense.date.desc()).all()

    r = 3; total = 0.0
    for e in expenses:
        data_cell(ws, r, 1, e.date.strftime("%d.%m.%Y") if e.date else "")
        data_cell(ws, r, 2, e.category or "—")
        data_cell(ws, r, 3, e.description or "—")
        data_cell(ws, r, 4, float(e.amount), number_format='#,##0', align=RIGHT, font=RED_FONT)
        data_cell(ws, r, 5, e.created_by or "—")
        ws.row_dimensions[r].height = 18
        total += float(e.amount); r += 1

    if r == 3:
        ws.merge_cells("A3:E3")
        ws.cell(3, 1).value = "Расходов нет"
        ws.cell(3, 1).font  = Font(name="Arial", size=11, color="888888")
        ws.cell(3, 1).alignment = CENTER; r = 4

    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r, 1).value = "ИТОГО"
    ws.cell(r, 1).font  = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    for col in range(1, 6):
        ws.cell(r, col).fill = SUBHEADER_FILL
        ws.cell(r, col).border = thin_border()
    data_cell(ws, r, 4, total, number_format='#,##0', align=RIGHT,
              font=Font(bold=True, name="Arial", size=10, color="C0392B"), fill=SUBHEADER_FILL)


# ─── Лист 6: Возвраты клиентам ───────────────────────────────────────────────

def build_returns_sheet(ws, db: Session, date_from: date, date_to: date):
    ws.title = "Возвраты клиентам"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"Возвраты клиентам за {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="E08030")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["Дата", "Клиент", "Товар", "Кол-во", "Сумма возврата", "Причина", "№ продажи"]
    widths = [16, 22, 30, 8, 18, 30, 12]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    returns = db.query(Return).join(Sale, Return.sale_id==Sale.id).filter(
        cast(Return.created_at, Date) >= date_from,
        cast(Return.created_at, Date) <= date_to,
    ).order_by(Return.created_at.desc()).all()

    r = 3; total = 0.0
    for ret in returns:
        sale     = db.query(Sale).filter(Sale.id == ret.sale_id).first()
        product  = db.query(Product).filter(Product.id == ret.product_id).first()
        cust_name = sale.customer.name if sale and sale.customer else "Розница"

        data_cell(ws, r, 1, ret.created_at.strftime("%d.%m.%Y %H:%M") if ret.created_at else "")
        data_cell(ws, r, 2, cust_name)
        data_cell(ws, r, 3, product.name if product else "—")
        data_cell(ws, r, 4, ret.quantity, align=CENTER)
        data_cell(ws, r, 5, float(ret.return_amount), number_format='#,##0', align=RIGHT, font=ORANGE_FONT)
        data_cell(ws, r, 6, ret.reason or "—")
        data_cell(ws, r, 7, ret.sale_id, align=CENTER)
        ws.row_dimensions[r].height = 18
        total += float(ret.return_amount); r += 1

    if r == 3:
        ws.merge_cells("A3:G3")
        ws.cell(3, 1).value = "Возвратов нет"
        ws.cell(3, 1).font  = Font(name="Arial", size=11, color="888888")
        ws.cell(3, 1).alignment = CENTER; r = 4

    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(r, 1).value = "ИТОГО ВОЗВРАТОВ"
    ws.cell(r, 1).font  = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    for col in range(1, 8):
        ws.cell(r, col).fill = SUBHEADER_FILL
        ws.cell(r, col).border = thin_border()
    data_cell(ws, r, 5, total, number_format='#,##0', align=RIGHT,
              font=Font(bold=True, name="Arial", size=10, color="E08030"), fill=SUBHEADER_FILL)


# ─── Лист 7: Возвраты поставщикам ────────────────────────────────────────────

def build_supplier_returns_sheet(ws, db: Session, date_from: date, date_to: date):
    ws.title = "Возвраты поставщикам"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Возвраты поставщикам за {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    t.font = Font(bold=True, size=13, name="Arial", color="7A3B8C")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    cols   = ["Дата", "Поставщик", "Товар", "Кол-во", "Цена", "Сумма", "Долг −", "Кредит +"]
    widths = [16, 22, 30, 8, 16, 18, 16, 14]
    header_row(ws, 2, cols, widths)
    ws.row_dimensions[2].height = 20

    sup_returns = db.query(SupplierReturn).filter(
        cast(SupplierReturn.created_at, Date) >= date_from,
        cast(SupplierReturn.created_at, Date) <= date_to,
    ).order_by(SupplierReturn.created_at.desc()).all()

    r = 3; total = 0.0
    for sr in sup_returns:
        sup = db.query(Supplier).filter(Supplier.id==sr.supplier_id).first()
        prod = db.query(Product).filter(Product.id==sr.product_id).first()

        data_cell(ws, r, 1, sr.created_at.strftime("%d.%m.%Y %H:%M") if sr.created_at else "")
        data_cell(ws, r, 2, sup.name if sup else "—")
        data_cell(ws, r, 3, prod.name if prod else "—")
        data_cell(ws, r, 4, sr.quantity, align=CENTER)
        data_cell(ws, r, 5, float(sr.purchase_price), number_format='#,##0', align=RIGHT)
        data_cell(ws, r, 6, float(sr.return_amount),  number_format='#,##0', align=RIGHT, font=ORANGE_FONT)
        data_cell(ws, r, 7, float(sr.debt_reduced),   number_format='#,##0', align=RIGHT,
                  font=Font(name="Arial", size=10, color="2481CC", bold=True) if float(sr.debt_reduced) > 0 else NORMAL_FONT)
        data_cell(ws, r, 8, float(sr.credit_added),   number_format='#,##0', align=RIGHT,
                  font=GREEN_FONT if float(sr.credit_added) > 0 else NORMAL_FONT)
        ws.row_dimensions[r].height = 18
        total += float(sr.return_amount); r += 1

    if r == 3:
        ws.merge_cells("A3:H3")
        ws.cell(3, 1).value = "Возвратов поставщикам нет"
        ws.cell(3, 1).font  = Font(name="Arial", size=11, color="888888")
        ws.cell(3, 1).alignment = CENTER; r = 4

    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(r, 1).value = "ИТОГО"
    ws.cell(r, 1).font  = SUBHEADER_FONT
    ws.cell(r, 1).alignment = CENTER
    for col in range(1, 9):
        ws.cell(r, col).fill = SUBHEADER_FILL
        ws.cell(r, col).border = thin_border()
    data_cell(ws, r, 6, total, number_format='#,##0', align=RIGHT,
              font=Font(bold=True, name="Arial", size=10, color="E08030"), fill=SUBHEADER_FILL)


# ─── Построить воркбук ────────────────────────────────────────────────────────

def build_workbook(db: Session, date_from: date, date_to: date) -> BytesIO:
    wb = openpyxl.Workbook()

    ws0 = wb.active
    build_summary_sheet(ws0, db, date_from, date_to)

    ws1 = wb.create_sheet()
    build_sales_sheet(ws1, db, date_from, date_to)

    ws2 = wb.create_sheet()
    build_stock_sheet(ws2, db)

    ws3 = wb.create_sheet()
    build_debts_sheet(ws3, db)

    ws4 = wb.create_sheet()
    build_suppliers_sheet(ws4, db)

    ws5 = wb.create_sheet()
    build_expenses_sheet(ws5, db, date_from, date_to)

    ws6 = wb.create_sheet()
    build_returns_sheet(ws6, db, date_from, date_to)

    ws7 = wb.create_sheet()
    build_supplier_returns_sheet(ws7, db, date_from, date_to)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("")
def export_excel(
    days: int = 30,
    alltime: bool = False,
    db: Session = Depends(get_db),
    _: int = Depends(get_current_user),
):
    date_to   = date.today()
    date_from = date(2000, 1, 1) if alltime else date_to - timedelta(days=days - 1)
    filename  = f"tradi_report_{date_to.strftime('%Y%m%d')}{'_alltime' if alltime else f'_{days}d'}.xlsx"

    buf = build_workbook(db, date_from, date_to)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/send")
def send_to_telegram(
    days: int = 30,
    alltime: bool = False,
    db: Session = Depends(get_db),
    telegram_id: int = Depends(get_current_user),
):
    """Отправляет Excel в Telegram-чат пользователя"""
    from app.core.config import settings

    user = db.query(User).filter(User.telegram_id == telegram_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    date_to   = date.today()
    date_from = date(2000, 1, 1) if alltime else date_to - timedelta(days=days - 1)
    period_label = "вся история" if alltime else f"{days} дней"
    filename  = f"tradi_report_{date_to.strftime('%Y%m%d')}{'_alltime' if alltime else ''}.xlsx"

    buf = build_workbook(db, date_from, date_to)

    from app.core.config import settings
    bot_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendDocument"

    with httpx.Client(timeout=60, verify=False) as client:
        response = client.post(
            bot_url,
            data={
                "chat_id": str(telegram_id),
                "caption": (
                    f"📊 Отчёт TRADI — {period_label}\n"
                    f"📅 {date_from.strftime('%d.%m.%Y') if not alltime else 'с начала'} — {date_to.strftime('%d.%m.%Y')}\n\n"
                    f"📋 Листы: Сводка · Продажи · Остатки · Долги · Поставщики · Расходы · Возвраты"
                ),
            },
            files={"document": (filename, buf.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    tg_result = response.json()
    if not tg_result.get("ok"):
        raise HTTPException(status_code=500,
            detail=f"Ошибка Telegram: {tg_result.get('description', 'неизвестная ошибка')}")

    return {"message": f"✅ Отчёт отправлен в Telegram! ({period_label})"}



# ─── Отправить SQL бэкап в Telegram ──────────────────────────────────────────

@router.get("/db-backup-send")
def send_backup_to_telegram(
    db: Session = Depends(get_db),
    telegram_id: int = Depends(get_current_user),
):
    """Создаёт SQL бэкап и отправляет в Telegram-чат пользователя"""
    from app.core.config import settings
    import psycopg2
    from urllib.parse import urlparse

    user = db.query(User).filter(User.telegram_id == telegram_id).first()
    if not user or user.role not in (UserRole.developer, UserRole.owner_business):
        raise HTTPException(status_code=403, detail="Только владелец или разработчик")

    # ── Генерируем SQL дамп ───────────────────────────────────────────────────
    try:
        parsed = urlparse(settings.DATABASE_URL)
        conn = psycopg2.connect(
            dbname=parsed.path.lstrip("/"), user=parsed.username,
            password=parsed.password, host=parsed.hostname, port=parsed.port or 5432,
        )
        conn.autocommit = True
        cur = conn.cursor()

        lines = [
            "-- Tradi backup\n",
            f"-- Date: {date.today()}\n\n",
            "SET client_encoding = 'UTF8';\n",
            "SET standard_conforming_strings = on;\n\n",
        ]

        cur.execute("SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename")
        tables = [r[0] for r in cur.fetchall()]

        for table in tables:
            lines.append(f"\n-- Table: {table}\n")
            lines.append(f"DELETE FROM {table};\n")
            cur.execute(f"SELECT * FROM {table}")
            rows = cur.fetchall()
            if not rows:
                continue
            cols = [desc[0] for desc in cur.description]
            cols_str = ", ".join(cols)
            for row in rows:
                values = []
                for v in row:
                    if v is None:             values.append("NULL")
                    elif isinstance(v, bool): values.append("TRUE" if v else "FALSE")
                    elif isinstance(v, (int, float)): values.append(str(v))
                    else:
                        escaped = str(v).replace("'", "''")
                        values.append(f"'{escaped}'")
                lines.append(f"INSERT INTO {table} ({cols_str}) VALUES ({', '.join(values)});\n")

        cur.execute("SELECT sequence_name FROM information_schema.sequences WHERE sequence_schema='public'")
        for (seq,) in cur.fetchall():
            cur.execute(f"SELECT last_value FROM {seq}")
            last_val = cur.fetchone()[0]
            lines.append(f"SELECT setval('{seq}', {last_val});\n")

        cur.close(); conn.close()
        content = "".join(lines).encode("utf-8")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка создания бэкапа: {e}")

    # ── Отправляем в Telegram ─────────────────────────────────────────────────
    filename = f"tradi_backup_{date.today().strftime('%Y%m%d')}.sql"
    bot_url  = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendDocument"

    with httpx.Client(timeout=60, verify=False) as client:
        response = client.post(
            bot_url,
            data={
                "chat_id": str(telegram_id),
                "caption": (
                    f"💾 Резервная копия базы данных TRADI\n"
                    f"📅 {date.today().strftime('%d.%m.%Y')}\n"
                    f"📦 Размер: {len(content)//1024} КБ\n\n"
                    f"Для восстановления загрузите этот файл обратно в приложение."
                ),
            },
            files={"document": (filename, content, "application/octet-stream")},
        )

    tg_result = response.json()
    if not tg_result.get("ok"):
        raise HTTPException(status_code=500,
            detail=f"Ошибка Telegram: {tg_result.get('description', 'неизвестная ошибка')}")

    return {"message": f"✅ Бэкап отправлен в Telegram! ({len(content)//1024} КБ)"}